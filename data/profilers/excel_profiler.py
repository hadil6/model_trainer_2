"""Excel Profiler — supports .xlsx, .xls, .xlsm (single and multi-sheet)."""
from __future__ import annotations

import io
from typing import Any

import openpyxl
import polars as pl

from core.config import get_settings
from data.profilers.base import BaseProfiler

_SENTINEL_VALUES = {"n/a", "na", "null", "none", "unknown", "-", "--", "?", "nan", ""}


class ExcelProfiler(BaseProfiler):
    def __init__(self) -> None:
        self.settings = get_settings()

    def _read_sheet(self, ws: Any) -> pl.DataFrame | None:
        rows = list(ws.iter_rows(values_only=True))
        while rows and all(v is None or str(v).strip() == "" for v in rows[0]):
            rows.pop(0)
        if len(rows) < 2:
            return None

        raw_headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        seen: dict[str, int] = {}
        headers: list[str] = []
        for h in raw_headers:
            if h in seen:
                seen[h] += 1
                headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                headers.append(h)

        data_rows = [
            list(row) for row in rows[1:]
            if not all(v is None or str(v).strip() == "" for v in row)
        ]
        if not data_rows:
            return None

        ncols = len(headers)
        data_rows = [row[:ncols] + [None] * max(0, ncols - len(row)) for row in data_rows]

        try:
            return pl.DataFrame(
                {h: [row[i] for row in data_rows] for i, h in enumerate(headers)},
                schema_overrides={},
                infer_schema_length=500,
            )
        except Exception:
            return pl.DataFrame(
                {h: [str(row[i]) if row[i] is not None else None for row in data_rows]
                 for i, h in enumerate(headers)}
            )

    @staticmethod
    def _col_stats(series: pl.Series) -> dict[str, Any]:
        row_count = series.len()
        null_count = series.null_count()
        null_ratio = null_count / max(row_count, 1)
        unique = series.n_unique()
        unique_ratio = unique / max(row_count, 1)

        numeric_stats: dict[str, Any] | None = None
        if series.dtype.is_numeric():
            non_null = series.drop_nulls()
            if non_null.len() > 0:
                numeric_stats = {
                    "min": non_null.min(),
                    "max": non_null.max(),
                    "mean": non_null.mean(),
                    "std": non_null.std(),
                    "q25": non_null.quantile(0.25),
                    "q50": non_null.quantile(0.50),
                    "q75": non_null.quantile(0.75),
                }

        outlier_count: int | None = None
        if series.dtype.is_numeric():
            non_null = series.drop_nulls()
            if non_null.len() >= 4:
                mean = non_null.mean() or 0
                std = non_null.std() or 0
                if std > 0:
                    outlier_count = int(((non_null - mean).abs() > 3 * std).sum())

        sentinel_count = 0
        if series.dtype == pl.Utf8:
            non_null = series.drop_nulls()
            lower = non_null.str.strip_chars().str.to_lowercase()
            sentinel_count = int(lower.is_in(list(_SENTINEL_VALUES)).sum())

        return {
            "name": series.name,
            "inferred_type": str(series.dtype),
            "null_ratio": round(float(null_ratio), 4),
            "unique_ratio": round(float(unique_ratio), 4),
            "numeric_stats": numeric_stats,
            "sentinel_count": sentinel_count,
            "outlier_count": outlier_count,
        }

    @staticmethod
    def _column_summary(df: pl.DataFrame, col_stats: list[dict]) -> list[dict[str, Any]]:
        summaries = []
        for cs in col_stats:
            s: dict[str, Any] = {
                "name": cs["name"],
                "type": cs["inferred_type"],
                "null_ratio": cs["null_ratio"],
            }
            if cs["numeric_stats"]:
                ns = cs["numeric_stats"]
                s["range"] = f"{ns['min']} - {ns['max']}"
                s["mean"] = round(ns["mean"], 2) if ns["mean"] is not None else None
                s["std"] = round(ns["std"], 2) if ns["std"] is not None else None
            else:
                try:
                    top = (
                        df[cs["name"]]
                        .drop_nulls()
                        .value_counts(sort=True)
                        .head(5)
                        .to_series(0)
                        .to_list()
                    )
                    s["top_values"] = [str(v) for v in top]
                except Exception:
                    pass
            summaries.append(s)
        return summaries

    @staticmethod
    def _clean_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        cleaned = []
        for row in rows:
            clean_row: dict[str, Any] = {}
            for k, v in row.items():
                if isinstance(v, str):
                    v = v.strip()
                    if v.lower() in _SENTINEL_VALUES:
                        v = None
                clean_row[k] = v
            cleaned.append(clean_row)
        return cleaned

    def _profile_sheet(self, df: pl.DataFrame, sheet_name: str, filename: str) -> dict[str, Any]:
        row_count = df.height
        col_count = df.width

        col_stats = [self._col_stats(df[col]) for col in df.columns]

        null_ratio_by_col = {cs["name"]: cs["null_ratio"] for cs in col_stats}
        unique_ratio_by_col = {cs["name"]: cs["unique_ratio"] for cs in col_stats}

        total_missing = sum(int(df[col].null_count()) for col in df.columns)
        dup_count = int(df.is_duplicated().sum()) if row_count else 0
        dup_ratio = dup_count / row_count if row_count else 0.0

        warnings: list[str] = []
        risk_flags: list[str] = []

        if row_count == 0:
            warnings.append("sheet_empty")
            risk_flags.append("empty_sheet")
        if row_count < 30:
            warnings.append(f"small_dataset_rows:{row_count}")
        if dup_ratio >= 0.10:
            risk_flags.append("high_duplicate_row_ratio")
            warnings.append(f"duplicate_rows:{dup_ratio:.4f}")
        for col_name, ratio in null_ratio_by_col.items():
            if ratio >= 0.10:
                warnings.append(f"missing_values:{col_name}:{ratio:.4f}")
            if ratio >= 0.30:
                risk_flags.append(f"high_missing_ratio:{col_name}")
        for col_name, ratio in unique_ratio_by_col.items():
            if ratio <= 0.01 and row_count >= 30:
                warnings.append(f"low_variance_column:{col_name}")
                risk_flags.append(f"near_constant_column:{col_name}")

        sample_size = min(300, row_count)
        sample_rows: list[dict] = []
        if sample_size > 0:
            sample_rows = df.sample(n=sample_size, seed=self.settings.sample_seed, shuffle=True).to_dicts()

        cleaned_rows = self._clean_rows(sample_rows)
        column_summaries = self._column_summary(df, col_stats)

        likely_id_cols = [
            cs["name"] for cs in col_stats
            if cs["unique_ratio"] >= 0.98 and cs["null_ratio"] <= 0.05
        ]
        target_candidates = sorted({
            cs["name"] for cs in col_stats
            if 1 < int(cs["unique_ratio"] * row_count) <= max(20, int(row_count * 0.2))
        })

        return {
            "sheet_name": sheet_name,
            "file_kind": "excel",
            "metadata": {
                "filename": filename,
                "sheet_name": sheet_name,
                "row_count": row_count,
                "column_count": col_count,
            },
            "quality_checks": {
                "total_missing_values": total_missing,
                "duplicate_row_count": dup_count,
                "duplicate_row_ratio": round(float(dup_ratio), 4),
                "high_null_columns": sorted(n for n, r in null_ratio_by_col.items() if r >= 0.30),
                "near_constant_columns": sorted(n for n, r in unique_ratio_by_col.items() if r <= 0.01 and row_count >= 30),
                "likely_id_columns": likely_id_cols,
                "possible_target_candidates": target_candidates,
            },
            "columns": col_stats,
            "column_summaries": column_summaries,
            "samples": {"rows": sample_rows, "cleaned_rows": cleaned_rows},
            "warnings": sorted(set(warnings)),
            "risk_flags": sorted(set(risk_flags)),
        }

    def profile(self, filename: str, content: bytes) -> dict[str, Any]:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        sheet_profiles: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df = self._read_sheet(ws)
            if df is None or df.height == 0:
                continue
            sheet_profiles.append(self._profile_sheet(df, sheet_name, filename))

        wb.close()

        if not sheet_profiles:
            return {
                "file_kind": "excel",
                "metadata": {"filename": filename, "size_bytes": len(content), "sheet_count": 0, "row_count": 0, "column_count": 0},
                "sheets": [],
                "samples": {"rows": [], "cleaned_rows": []},
                "columns": [],
                "column_summaries": [],
                "quality_checks": {},
                "warnings": ["excel_no_readable_sheets"],
                "risk_flags": ["empty_file"],
            }

        primary = max(sheet_profiles, key=lambda s: s["metadata"]["row_count"])
        total_rows = sum(s["metadata"]["row_count"] for s in sheet_profiles)
        all_warnings = sorted({w for s in sheet_profiles for w in s["warnings"]})
        all_risk_flags = sorted({r for s in sheet_profiles for r in s["risk_flags"]})

        merged_cleaned: list[dict] = []
        for s in sheet_profiles:
            for row in s["samples"]["cleaned_rows"]:
                merged_cleaned.append({"_sheet": s["metadata"]["sheet_name"], **row})

        merged_raw: list[dict] = []
        for s in sheet_profiles:
            for row in s["samples"]["rows"]:
                merged_raw.append({"_sheet": s["metadata"]["sheet_name"], **row})

        return {
            "file_kind": "excel",
            "metadata": {
                "filename": filename,
                "size_bytes": len(content),
                "sheet_count": len(sheet_profiles),
                "sheet_names": [s["metadata"]["sheet_name"] for s in sheet_profiles],
                "row_count": total_rows,
                "column_count": primary["metadata"]["column_count"],
            },
            "quality_checks": primary["quality_checks"],
            "columns": primary["columns"],
            "column_summaries": primary["column_summaries"],
            "samples": {"rows": merged_raw[:300], "cleaned_rows": merged_cleaned[:300]},
            "sheets": sheet_profiles,
            "warnings": all_warnings,
            "risk_flags": all_risk_flags,
        }
